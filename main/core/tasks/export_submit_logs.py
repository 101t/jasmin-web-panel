import csv
import io
from datetime import datetime
from celery import shared_task
from celery.utils.log import get_task_logger
from django.core.cache import cache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from main.core.models import SubmitLog

logger = get_task_logger(__name__)


@shared_task(bind=True)
def export_submit_logs_task(self, filters, export_format='csv'):
    """
    Export submit logs to CSV or Excel format.
    
    Args:
        filters: Dictionary containing filter parameters
        export_format: 'csv' or 'xlsx'
    
    Returns:
        Dictionary with status and file path or error message
    """
    task_id = None
    try:
        task_id = self.request.id
        logger.info(f"=== TASK STARTED === ID: {task_id}")
        logger.info(f"Export format: {export_format}")
        logger.info(f"Filters: {filters}")
        
        # Test database connection
        logger.info("Testing database connection...")
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            logger.info("Database connection OK")
        
        # Update progress: Starting
        logger.info("Setting initial cache progress...")
        cache.set(f'export_progress_{task_id}', {'status': 'processing', 'progress': 0}, 300)
        logger.info("Cache progress set successfully")
        
        # Build queryset based on filters
        logger.info("Building queryset...")
        queryset = SubmitLog.objects.all()
        logger.info(f"Initial queryset created")
        
        # Apply date range filters
        if filters.get('date_column') and filters.get('date_from'):
            date_column = filters['date_column']  # 'created_at' or 'status_at'
            date_from = datetime.fromisoformat(filters['date_from'])
            queryset = queryset.filter(**{f'{date_column}__gte': date_from})
        
        if filters.get('date_column') and filters.get('date_to'):
            date_column = filters['date_column']
            date_to = datetime.fromisoformat(filters['date_to'])
            queryset = queryset.filter(**{f'{date_column}__lte': date_to})
        
        # Apply search filter
        if filters.get('search'):
            from django.db.models import Q
            search_query = filters['search']
            queryset = queryset.filter(
                Q(msgid__icontains=search_query) |
                Q(source_addr__icontains=search_query) |
                Q(destination_addr__icontains=search_query) |
                Q(short_message__icontains=search_query) |
                Q(uid__icontains=search_query)
            )
        
        # Apply status filter
        if filters.get('status_filter') == 'success':
            queryset = queryset.filter(status__in=['ESME_ROK', 'ESME_RINVNUMDESTS'])
        elif filters.get('status_filter') == 'fail':
            queryset = queryset.filter(status='ESME_RDELIVERYFAILURE')
        elif filters.get('status_filter') == 'unknown':
            queryset = queryset.exclude(status__in=['ESME_ROK', 'ESME_RINVNUMDESTS', 'ESME_RDELIVERYFAILURE'])
        
        # Order by created_at
        logger.info("Ordering queryset by created_at...")
        queryset = queryset.order_by('-created_at')
        
        logger.info("Counting total records...")
        total_records = queryset.count()
        logger.info(f"Total records to export: {total_records}")
        
        if total_records == 0:
            cache.set(f'export_progress_{task_id}', {
                'status': 'completed',
                'progress': 100,
                'error': 'No records found with the given filters'
            }, 300)
            return {'status': 'error', 'message': 'No records found'}
        
        # Update progress: Fetching data
        cache.set(f'export_progress_{task_id}', {'status': 'processing', 'progress': 10}, 300)
        
        # Define headers
        headers = [
            'Message ID', 'Source Connector', 'Routed CID', 'Source Address',
            'Destination Address', 'Rate', 'PDU Count', 'Short Message',
            'Status', 'UID', 'Trials', 'Created At', 'Status At'
        ]
        
        if export_format == 'csv':
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            
            # Write data in batches
            batch_size = 1000
            processed = 0
            
            for record in queryset.iterator(chunk_size=batch_size):
                writer.writerow([
                    record.msgid,
                    record.source_connector,
                    record.routed_cid,
                    record.decoded_source_addr,
                    record.decoded_destination_addr,
                    str(record.rate),
                    record.pdu_count,
                    record.decoded_short_message,
                    record.status,
                    record.uid,
                    record.trials,
                    record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else '',
                    record.status_at.strftime('%Y-%m-%d %H:%M:%S') if record.status_at else '',
                ])
                
                processed += 1
                if processed % 100 == 0:
                    progress = 10 + int((processed / total_records) * 80)
                    cache.set(f'export_progress_{task_id}', {
                        'status': 'processing',
                        'progress': progress,
                        'processed': processed,
                        'total': total_records
                    }, 300)
            
            content = output.getvalue()
            content_type = 'text/csv'
            filename = f'submit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            
        else:  # xlsx
            # Generate Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Submit Logs"
            
            # Style for header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Write data in batches
            batch_size = 1000
            processed = 0
            row_num = 2
            
            for record in queryset.iterator(chunk_size=batch_size):
                ws.cell(row=row_num, column=1, value=record.msgid)
                ws.cell(row=row_num, column=2, value=record.source_connector)
                ws.cell(row=row_num, column=3, value=record.routed_cid)
                ws.cell(row=row_num, column=4, value=record.decoded_source_addr)
                ws.cell(row=row_num, column=5, value=record.decoded_destination_addr)
                ws.cell(row=row_num, column=6, value=float(record.rate) if record.rate else 0.0)
                ws.cell(row=row_num, column=7, value=record.pdu_count)
                ws.cell(row=row_num, column=8, value=record.decoded_short_message)
                ws.cell(row=row_num, column=9, value=record.status)
                ws.cell(row=row_num, column=10, value=record.uid)
                ws.cell(row=row_num, column=11, value=record.trials)
                ws.cell(row=row_num, column=12, value=record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else '')
                ws.cell(row=row_num, column=13, value=record.status_at.strftime('%Y-%m-%d %H:%M:%S') if record.status_at else '')
                
                row_num += 1
                processed += 1
                
                if processed % 100 == 0:
                    progress = 10 + int((processed / total_records) * 80)
                    cache.set(f'export_progress_{task_id}', {
                        'status': 'processing',
                        'progress': progress,
                        'processed': processed,
                        'total': total_records
                    }, 300)
            
            # Auto-size columns
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'submit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Store the file content in cache temporarily (for 5 minutes)
        cache.set(f'export_file_{task_id}', {
            'content': content,
            'content_type': content_type,
            'filename': filename
        }, 300)
        
        # Update progress: Completed
        cache.set(f'export_progress_{task_id}', {
            'status': 'completed',
            'progress': 100,
            'total': total_records,
            'filename': filename
        }, 300)
        
        logger.info(f"Export completed: {filename} ({total_records} records)")
        
        return {
            'status': 'success',
            'filename': filename,
            'total_records': total_records
        }
        
    except Exception as e:
        logger.error(f"Export failed with error: {str(e)}", exc_info=True)
        if task_id:
            cache.set(f'export_progress_{task_id}', {
                'status': 'failed',
                'progress': 0,
                'error': str(e)
            }, 300)
        return {'status': 'error', 'message': str(e)}
